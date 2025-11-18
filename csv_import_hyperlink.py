#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Extract hyperlinks from a given column in an Excel .xlsx file and save to CSV.

Usage examples:
  python csv_import_hyperlink.py "case_ratings.xlsx" "case_ratings_with_urls.csv" --col "Title" --sheet "case_ratings"
  python csv_import_hyperlink.py "case_ratings.xlsx" "case_ratings_with_urls.csv" --col D
  python csv_import_hyperlink.py "case_ratings.xlsx" "case_ratings_with_urls.csv" --col 4

Notes:
- Works with .xlsx only (openpyxl). If you have .xls, open it in Excel and Save As .xlsx.
- The --col argument can be a header name, an Excel column letter (A, B, C...), or a 1-based index.
"""

import argparse
import os
import sys
import re
from typing import Optional, Tuple, List, Any, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def eprint(*args, **kwargs):
    print(*args, file=sys.stderr, **kwargs)


def normalize_col_ref(col_ref: str) -> Tuple[str, Optional[int]]:
    """
    Normalize a user-given column reference.

    Returns:
        ("name", None)                if user provided a header name
        ("letter", zero_based_index)  if user provided column letter
        ("index", zero_based_index)   if user provided numeric index (1-based in Excel)
    """
    s = str(col_ref).strip()

    # Numeric index? (1-based)
    if s.isdigit():
        idx = int(s) - 1
        if idx < 0:
            raise ValueError("Column index must be >= 1.")
        return ("index", idx)

    # Column letter? (A, B, ... AA, AB ...)
    letters = s.replace(" ", "")
    if letters and letters.isalpha() and letters.upper() == letters:
        def letters_to_index(letters_: str) -> int:
            acc = 0
            for ch in letters_:
                if not ("A" <= ch <= "Z"):
                    raise ValueError(f"Invalid column letter: {letters}")
                acc = acc * 26 + (ord(ch) - ord("A") + 1)
            return acc - 1  # zero-based
        return ("letter", letters_to_index(letters.upper()))

    # Otherwise treat as a header name
    return ("name", None)


def find_target_column_index(ws: Worksheet, mode: str, idx: Optional[int], header_name: Optional[str]) -> int:
    """
    Determine the zero-based column index in the worksheet for the target column.

    If mode == "index" or "letter": idx is already zero-based.
    If mode == "name": use the first row as header row to find header_name.
    """
    if mode in ("index", "letter"):
        if idx is None or idx < 0:
            raise ValueError("Internal error: invalid zero-based column index.")
        return idx

    # mode == "name": scan first row for matching header
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False), None)
    if header_row is None:
        raise ValueError("Worksheet appears to be empty (no header row).")

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]

    # Exact match first
    if header_name in headers:
        return headers.index(header_name)

    # Case-insensitive match
    lowered = [h.lower() for h in headers]
    if header_name is not None and header_name.lower() in lowered:
        return lowered.index(header_name.lower())

    raise ValueError(
        f"Column header '{header_name}' not found in first row. "
        f"Available headers: {headers}"
    )


# Regex to parse =HYPERLINK("url","text") or =HYPERLINK("url";"text") (some locales use ;)
HYPERLINK_RE = re.compile(
    r'^\s*=\s*HYPERLINK\s*\(\s*"([^"]+)"\s*[;,]',
    flags=re.IGNORECASE
)


def extract_url_from_cell(ws: Worksheet, cell) -> Optional[str]:
    """
    Try multiple ways to extract a hyperlink from a cell:
    1) cell.hyperlink.target (explicit hyperlinks)
    2) HYPERLINK("url","text") formula
    3) worksheet._hyperlinks mapping (style-applied links)
    """
    if cell is None:
        return None

    # 1) Explicit hyperlink object
    link_obj = getattr(cell, "hyperlink", None)
    if link_obj is not None and getattr(link_obj, "target", None):
        return link_obj.target

    # 2) HYPERLINK formula (data_only=False to see formulas)
    if isinstance(cell.value, str) and cell.value.startswith("="):
        m = HYPERLINK_RE.match(cell.value)
        if m:
            return m.group(1)

    # 3) Private hyperlink registry on the sheet (style-applied hyperlinks)
    #    e.g., users clicked "Insert Hyperlink..." on displayed text
    #    openpyxl keeps these in ws._hyperlinks with .ref (cell range) and .target
    try:
        for hl in getattr(ws, "_hyperlinks", []):
            # hl.ref can be a range ("D2:D200") or a single cell address
            # simplest: if our cell coordinate appears in that ref string
            if cell.coordinate in str(hl.ref):
                if getattr(hl, "target", None):
                    return hl.target
    except Exception:
        pass

    return None


def worksheet_to_dataframe_with_urls(ws: Worksheet, target_col_idx: int, url_col_name: str) -> pd.DataFrame:
    """
    Read the worksheet into a DataFrame, preserving values, and add a URL column
    extracted from the hyperlink of the target column cells.

    Assumes the first row is header.
    """
    rows: List[Dict[str, Any]] = []

    # Header
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
    num_cols = len(headers)
    if num_cols == 0:
        raise ValueError("No headers found in the first row; cannot proceed.")

    # Data rows
    for r_cells in ws.iter_rows(min_row=2, values_only=False):
        cells = list(r_cells)
        if len(cells) < num_cols:
            cells += [None] * (num_cols - len(cells))

        row_dict: Dict[str, Any] = {}
        for j in range(num_cols):
            cell = cells[j]
            row_dict[headers[j]] = cell.value if cell is not None else None

        # URL from the target column
        url_val = None
        if target_col_idx < len(cells):
            tcell = cells[target_col_idx]
            url_val = extract_url_from_cell(ws, tcell)

        row_dict[url_col_name] = url_val
        rows.append(row_dict)

    df = pd.DataFrame(rows, columns=headers + [url_col_name])
    return df


def main():
    ap = argparse.ArgumentParser(
        description="Extract hyperlinks from a given column in an Excel .xlsx file and write CSV."
    )
    ap.add_argument("input", help="Path to input Excel file (.xlsx)")
    ap.add_argument("output_csv", help="Path to output CSV file")
    ap.add_argument("--sheet", help="Worksheet name (optional; defaults to the first sheet)", default=None)
    ap.add_argument("--col", required=True,
                    help="Target column (header name, Excel letter like A/D/AA, or 1-based numeric index)")
    ap.add_argument("--url_col_name", default="URL", help="Name of the output URL column (default: URL)")
    args = ap.parse_args()

    in_path = args.input
    out_path = args.output_csv
    sheet_name = args.sheet
    col_ref = args.col
    url_col_name = args.url_col_name

    if not os.path.exists(in_path):
        eprint(f"[ERROR] Input file not found: {in_path}")
        sys.exit(1)

    _, ext = os.path.splitext(in_path)
    if ext.lower() != ".xlsx":
        eprint("[WARNING] This script supports .xlsx only. If you have .xls, open in Excel and Save As .xlsx.")

    # IMPORTANT: data_only=False so we can see HYPERLINK formulas
    try:
        wb = load_workbook(in_path, data_only=False, read_only=False)
    except Exception as e:
        eprint(f"[ERROR] Failed to open workbook '{in_path}': {e}")
        sys.exit(1)

    sheetnames = wb.sheetnames or []
    if not sheetnames:
        eprint(f"[ERROR] No sheets found in workbook '{in_path}'.")
        sys.exit(1)

    if sheet_name:
        if sheet_name not in sheetnames:
            eprint(f"[ERROR] Sheet '{sheet_name}' not found. Available: {sheetnames}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb[sheetnames[0]]
        eprint(f"[INFO] --sheet not given. Using first sheet: '{ws.title}'")

    mode, idx = normalize_col_ref(col_ref)
    try:
        target_idx = find_target_column_index(ws, mode, idx, None if mode != "name" else col_ref)
    except Exception as e:
        eprint(f"[ERROR] {e}")
        sys.exit(1)

    try:
        df = worksheet_to_dataframe_with_urls(ws, target_idx, url_col_name=url_col_name)
    except Exception as e:
        eprint(f"[ERROR] Could not build DataFrame from worksheet '{ws.title}': {e}")
        sys.exit(1)

    try:
        df.to_csv(out_path, index=False)
    except Exception as e:
        eprint(f"[ERROR] Failed to write CSV to '{out_path}': {e}")
        sys.exit(1)

    # Stats
    found = df[url_col_name].notna().sum() if url_col_name in df.columns else 0
    print(f"[OK] Wrote {len(df)} rows to: {out_path}")
    print(f"[OK] Extracted URLs in column '{url_col_name}': {found}/{len(df)}")


if __name__ == "__main__":
    main()